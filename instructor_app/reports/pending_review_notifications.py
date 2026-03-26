import datetime
from io import BytesIO

from django import forms
from django.urls import reverse_lazy
from django.core.files.base import ContentFile

import openpyxl
from openpyxl.utils import get_column_letter

from crispy_forms.helper import FormHelper
from crispy_forms.layout import Submit

from cis.backends.storage_backend import PrivateMediaStorage

from ..services.pending_review_notifications import get_pending_review_notifications


class pending_review_notifications(forms.Form):
    """
    Export the list of faculty reviewers who have an outstanding course
    application review (status '---').

    No filter options — the result reflects the live state of reviewer
    assignments.
    """

    roles = []
    request = None

    def __init__(self, request=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.request = request

        self.helper = FormHelper()
        self.helper.attrs = {'target': '_blank'}
        self.helper.form_method = 'POST'
        self.helper.add_input(Submit('submit', 'Generate Export'))

        if self.request:
            self.helper.form_action = reverse_lazy(
                'report:run_report', args=[request.GET.get('report_id')]
            )

    def run(self, task, data):
        pending = get_pending_review_notifications()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pending Review Notifications'

        headers = [
            'Reviewer Last Name', 'Reviewer First Name', 'Reviewer Email',
            'Course', 'Applicant', 'Assigned On',
        ]
        ws.append(headers)

        for entry in pending:
            r = entry['reviewer']
            ws.append([
                r.reviewer.last_name,
                r.reviewer.first_name,
                entry['reviewer_email'],
                str(entry['course']),
                entry['teacher_name'],
                entry['assigned_on'].strftime('%m/%d/%Y') if entry['assigned_on'] else '',
            ])

        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value)) for cell in ws[col_letter] if cell.value),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file_name = 'pending_review_notifications-' + str(datetime.datetime.now()) + '.xlsx'
        path = 'reports/' + str(task.id) + '/' + file_name
        media_storage = PrivateMediaStorage()
        path = media_storage.save(path, ContentFile(buffer.read()))
        path = media_storage.url(path)

        return path
