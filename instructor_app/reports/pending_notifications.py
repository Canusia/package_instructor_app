import datetime

from django import forms
from django.urls import reverse_lazy
from django.core.files.base import ContentFile

import openpyxl
from openpyxl.utils import get_column_letter

from crispy_forms.helper import FormHelper
from crispy_forms.layout import Submit

from cis.backends.storage_backend import PrivateMediaStorage

from ..services.incomplete_notifications import get_pending_notifications


class pending_notifications(forms.Form):
    """
    Export a list of applicants who would receive an incomplete-application
    notification if the cron job ran right now.

    No filter options — the result is driven entirely by the frequency and
    missing-items settings configured in incomplete_si_application and
    inst_app_language.
    """

    roles = []
    request = None

    def __init__(self, request=None, *args, **kwargs):
        super().__init__(*args, **kwargs)

        self.request = request

        self.helper = FormHelper()
        self.helper.attrs = {'target': '_blank'}
        self.helper.form_method = 'POST'
        self.helper.add_input(Submit('submit', 'Generate Export'))

        if self.request:
            self.helper.form_action = reverse_lazy(
                'report:run_report', args=[request.GET.get('report_id')]
            )

    def run(self, task, data):
        pending = get_pending_notifications()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Pending Notifications'

        headers = [
            'Last Name', 'First Name', 'Email', 'Send To',
            'Missing Items', 'Recommendation Reminders',
        ]
        ws.append(headers)

        if isinstance(pending, str):
            # Feature inactive — write the skip reason as a single row
            ws.append([pending])
        else:
            for entry in pending:
                app = entry['app']
                ws.append([
                    app.user.last_name,
                    app.user.first_name,
                    app.user.email,
                    entry['to_email'],
                    '; '.join(entry['missing_items']),
                    '; '.join(
                        f"{r['name']} <{r['email']}>"
                        for r in entry['recommenders_to_remind']
                    ),
                ])

        # Auto-size columns
        for col_idx, _ in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                (len(str(cell.value)) for cell in ws[col_letter] if cell.value),
                default=10,
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

        from io import BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        file_name = 'pending_notifications-' + str(datetime.datetime.now()) + '.xlsx'
        path = 'reports/' + str(task.id) + '/' + file_name
        media_storage = PrivateMediaStorage()
        path = media_storage.save(path, ContentFile(buffer.read()))
        path = media_storage.url(path)

        return path
